"""
Peer Comparison Excel Report

Sprint 3 - Day 20
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from openpyxl.utils.dataframe import dataframe_to_rows
# Metrics to include in the report
REPORT_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
    "composite_quality_score",
]

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
)
AMBER_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD966",
)

def prepare_peer_report_dataframe(
    analytics_df: pd.DataFrame,
    peer_percentiles_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Merge analytics data with peer percentile rankings.
    """

    percentile_df = (
        peer_percentiles_df
        .pivot(
            index=["company_id", "year"],
            columns="metric",
            values="percentile_rank",
        )
        .reset_index()
    )

    percentile_df.columns.name = None
    percentile_columns = {
    column: f"{column}_percentile"
    for column in percentile_df.columns
        if column not in ["company_id", "year"]
}

    percentile_df = percentile_df.rename(columns=percentile_columns)
    report_df = analytics_df.merge(
        percentile_df,
        on=["company_id", "year"],
        how="left",
    )

    return report_df

def create_peer_group_sheet(
    workbook: Workbook,
    peer_group_df: pd.DataFrame,
    sheet_name: str,
):
    """
    Create one worksheet for a peer group.
    """

    worksheet = workbook.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(
        peer_group_df,
        index=False,
        header=True,
    ):
        worksheet.append(row)

    return worksheet

def apply_percentile_formatting(
    worksheet,
):
    """
    Apply color formatting to percentile columns.
    """

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    percentile_columns = []

    for index, column in enumerate(headers, start=1):
        if (
            isinstance(column, str)
            and column.endswith("_percentile")
        ):
            percentile_columns.append(index)

    for column in percentile_columns:

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            value = worksheet.cell(
                row=row,
                column=column,
            ).value

            if value is None:
                continue

            if value >= 75:
                worksheet.cell(
                    row=row,
                    column=column,
                ).fill = GREEN_FILL

            elif value <= 25:
                worksheet.cell(
                    row=row,
                    column=column,
                ).fill = RED_FILL

            else:
                worksheet.cell(
                    row=row,
                    column=column,
                ).fill = YELLOW_FILL

def highlight_benchmark_company(
    worksheet,
):
    """
    Highlight the company with the highest
    composite quality score.
    """

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    score_column = headers.index(
        "composite_quality_score"
    ) + 1

    highest_score = -1
    benchmark_row = None

    for row in range(
        2,
        worksheet.max_row + 1,
    ):

        value = worksheet.cell(
            row=row,
            column=score_column,
        ).value

        if value is not None and value > highest_score:
            highest_score = value
            benchmark_row = row

    if benchmark_row is None:
        return

    for column in range(
        1,
        worksheet.max_column + 1,
    ):
        worksheet.cell(
            row=benchmark_row,
            column=column,
        ).fill = AMBER_FILL

def add_median_row(
    worksheet,
):
    """
    Add a median summary row
    at the bottom of the worksheet.
    """

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    median_row = ["Median"]

    for column in headers[1:]:

        values = []

        column_index = headers.index(column) + 1

        for row in range(2, worksheet.max_row + 1):

            value = worksheet.cell(
                row=row,
                column=column_index,
            ).value

            if isinstance(value, (int, float)):
                values.append(value)

        if values:
            median_row.append(pd.Series(values).median())
        else:
            median_row.append("")

    worksheet.append(median_row)

from pathlib import Path


def export_peer_comparison_report(
    report_df: pd.DataFrame,
    output_path: str = "output/peer_comparison.xlsx",
):
    """
    Export peer comparison report to Excel.
    """

    workbook = Workbook()

    # Remove the default sheet
    workbook.remove(workbook.active)

    peer_groups = sorted(
        report_df["peer_group_name"]
        .dropna()
        .unique()
    )

    for peer_group in peer_groups:

        peer_group_df = (
            report_df[
                report_df["peer_group_name"] == peer_group
            ]
            .copy()
        )

        worksheet = create_peer_group_sheet(
            workbook,
            peer_group_df,
            peer_group,
        )

        apply_percentile_formatting(
            worksheet,
        )

        highlight_benchmark_company(
            worksheet,
        )

        add_median_row(
            worksheet,
        )

    Path(output_path).parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_path)

    return output_path