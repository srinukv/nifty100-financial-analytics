from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
OUTPUT_FILE = Path("output/screener_output.xlsx")
EXPORT_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "sub_sector",
    "composite_quality_score",
    "sector_relative_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "cfo_pat_ratio",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "sales",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
]
GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)
def create_output_directory():
    """
    Create the output directory if it does not exist.
    """

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

def export_screeners_to_excel(
    screener_results: dict,
):
    """
    Export all screener results to an Excel workbook.
    """

    create_output_directory()

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        for sheet_name, df in screener_results.items():
            df = df.sort_values(
    by="composite_quality_score",
    ascending=False,
)
            df = df[EXPORT_COLUMNS]
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    apply_conditional_formatting()
def apply_conditional_formatting():
    """
    Apply preset-specific conditional formatting.
    """

    workbook = load_workbook(OUTPUT_FILE)

    PRESET_RULES = {
        "quality_compounder": {
            "return_on_equity_pct": (">", 15),
            "debt_to_equity": ("<", 1.0),
            "free_cash_flow_cr": (">", 0),
            "revenue_cagr_5yr": (">", 10),
        },
        "growth_accelerator": {
            "pat_cagr_5yr": (">", 20),
            "revenue_cagr_5yr": (">", 15),
            "debt_to_equity": ("<", 2.0),
        },
        "debt_free_blue_chip": {
            "debt_to_equity": ("==", 0),
            "return_on_equity_pct": (">", 12),
            "sales": (">", 5000),
        },
    }

    for sheet in workbook.worksheets:

        if sheet.title not in PRESET_RULES:
            continue

        rules = PRESET_RULES[sheet.title]

        headers = {
            cell.value: cell.column
            for cell in sheet[1]
        }

        for metric, (operator, threshold) in rules.items():

            if metric not in headers:
                continue

            column = headers[metric]

            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(row=row, column=column)

                value = cell.value

                if value is None:
                    continue

                passed = False

                if operator == ">":
                    passed = value > threshold

                elif operator == "<":
                    passed = value < threshold

                elif operator == "==":
                    passed = value == threshold

                if passed:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL

    workbook.save(OUTPUT_FILE)