import pandas as pd

from src.reports.peer_report import prepare_peer_report_dataframe
from openpyxl import Workbook
from src.reports.peer_report import create_peer_group_sheet
from src.reports.peer_report import apply_percentile_formatting
from src.reports.peer_report import highlight_benchmark_company
from src.reports.peer_report import add_median_row

def test_prepare_peer_report_dataframe():
    """
    Test merging analytics data with peer percentile data.
    """

    analytics_df = pd.DataFrame({
        "company_id": ["ABC"],
        "company_name": ["ABC Ltd"],
        "year": [2024],
        "return_on_equity_pct": [20],
    })

    peer_percentiles_df = pd.DataFrame({
        "company_id": ["ABC"],
        "year": [2024],
        "metric": ["return_on_equity_pct"],
        "percentile_rank": [92],
    })

    report_df = prepare_peer_report_dataframe(
        analytics_df,
        peer_percentiles_df,
    )

    assert report_df.shape[0] == 1
    assert report_df.iloc[0]["company_name"] == "ABC Ltd"
    assert report_df.iloc[0]["return_on_equity_pct"] == 20
    assert report_df.iloc[0]["return_on_equity_pct_percentile"] == 92

def test_create_peer_group_sheet():
    """
    Test worksheet creation for one peer group.
    """

    workbook = Workbook()

    # Remove the default sheet created by Workbook()
    workbook.remove(workbook.active)

    peer_group_df = pd.DataFrame({
        "company_id": ["ABC"],
        "company_name": ["ABC Ltd"],
        "return_on_equity_pct": [20],
        "return_on_equity_pct_percentile": [92],
    })

    worksheet = create_peer_group_sheet(
        workbook,
        peer_group_df,
        "IT",
    )

    assert worksheet.title == "IT"

    # Header row
    assert worksheet["A1"].value == "company_id"
    assert worksheet["B1"].value == "company_name"

    # First data row
    assert worksheet["A2"].value == "ABC"
    assert worksheet["B2"].value == "ABC Ltd"

def test_apply_percentile_formatting():
    """
    Test conditional formatting for percentile columns.
    """

    workbook = Workbook()

    workbook.remove(workbook.active)

    peer_group_df = pd.DataFrame({
        "company_id": ["A", "B", "C"],
        "return_on_equity_pct_percentile": [90, 50, 20],
    })

    worksheet = create_peer_group_sheet(
        workbook,
        peer_group_df,
        "IT",
    )

    apply_percentile_formatting(worksheet)

    # Green
    assert worksheet["B2"].fill.start_color.rgb.endswith("C6EFCE")

    # Yellow
    assert worksheet["B3"].fill.start_color.rgb.endswith("FFEB9C")

    # Red
    assert worksheet["B4"].fill.start_color.rgb.endswith("FFC7CE")


def test_highlight_benchmark_company():
    """
    Test that the highest composite quality score
    row is highlighted.
    """

    workbook = Workbook()

    workbook.remove(workbook.active)

    peer_group_df = pd.DataFrame({
        "company_id": ["A", "B", "C"],
        "company_name": ["A Ltd", "B Ltd", "C Ltd"],
        "composite_quality_score": [70, 95, 82],
    })

    worksheet = create_peer_group_sheet(
        workbook,
        peer_group_df,
        "IT",
    )

    highlight_benchmark_company(worksheet)

    # Row 3 contains the highest score (95)
    assert worksheet["A3"].fill.start_color.rgb.endswith("FFD966")
    assert worksheet["B3"].fill.start_color.rgb.endswith("FFD966")
    assert worksheet["C3"].fill.start_color.rgb.endswith("FFD966")

def test_add_median_row():
    """
    Test that the median row is added correctly.
    """

    workbook = Workbook()

    workbook.remove(workbook.active)

    peer_group_df = pd.DataFrame({
        "company_id": ["A", "B", "C"],
        "company_name": ["A Ltd", "B Ltd", "C Ltd"],
        "return_on_equity_pct": [10, 20, 30],
        "composite_quality_score": [70, 80, 90],
    })

    worksheet = create_peer_group_sheet(
        workbook,
        peer_group_df,
        "IT",
    )

    add_median_row(worksheet)

    median_row = worksheet.max_row

    assert worksheet.cell(row=median_row, column=1).value == "Median"

    # Median of [10, 20, 30]
    assert worksheet.cell(row=median_row, column=3).value == 20

    # Median of [70, 80, 90]
    assert worksheet.cell(row=median_row, column=4).value == 80